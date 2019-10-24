import openpyxl
import sys
import os


caminho = 'D:\\CompartilhaVM\\lusca.xlsx'
#arq1 = 'D:\\CompartilhaVM\\testeArquivo.png'
arq1 = 'D:\\CompartilhaVM\\images.png'

#workbook
wb = openpyxl.load_workbook(caminho)
#sheet
ws = wb.worksheets[0]
nomeSheet = wb.sheetnames[0]

#recebe a imagem.
img = openpyxl.drawing.image.Image(arq1)
#posiciona imagem na coluna C linha 3.
img.anchor = 'C3'

#add imagem
ws.add_image(img)
#salva nova planilha devido ao nome ser diferente. Com a imagem.
wb.save('D:\\CompartilhaVM\\out.xlsx')


#get
#sheet = wb.get_sheet_by_name('Plan1')

#le os valores das celulas.
a1 = ws['A1'].value
a2 = ws['A2'].value
b1 = ws['B1'].value
b2 = ws['B2'].value

print(ws)
print(a1, a2, b1, b2)
print(nomeSheet)
#print(nameImg)
